# -*- cocoding:utf-8 -*-
from openpyxl import load_workbook


class OpenpyxlUtil:
    def __init__(self, excel_path, sheet_name):
        # 加载Excel数据
        # self.wb=openpyxl.Workbook(excel_path)
        # PS：Workbook和load_workbook相同，返回的都是一个Workbook对象。
        self.wb = load_workbook(excel_path)
        # 获取所有的sheet
        print(self.wb.sheetnames)
        # 获取当前活跃工作簿(关闭文档时的sheet或当前打开的sheet)
        print(self.wb.active)
        # 判断是否以只读模式打开Excel
        print(self.wb.read_only)
        # 获取文档的字符集编码
        print(self.wb.encoding)
        # 获取文档的元数据，如标题，创建者，创建日期等
        print(self.wb.properties)
        # 创建一个空的sheet
        self.wb.create_sheet()
        # 加载sheet_name工作表数据
        self.sheet = self.wb[sheet_name]

    def getDataFromSheet(self):
        data_list = []
        # 获取表格的标题
        title_name = self.sheet.title
        # 获取sheet_name工作表中的最大行号
        max_row_num = self.sheet.max_row
        # 获取sheet_name工作表中的最小行号
        min_row_num = self.sheet.min_row
        # 获取表格的最大列
        max_column_num = self.sheet.max_column
        # 获取表格的最小列
        min_column_num = self.sheet.min_column
        # 按行获取单元格(Cell对象) - 生成器
        rows_Iterator = self.sheet.rows
        # columns：按列获取单元格(Cell对象) - 生成器
        columns_Iterator = self.sheet.columns
        self.sheet.iter_rows  # 按行获取所有单元格，内置属性有(min_row,max_row,min_col,max_col)
        self.sheet.append("星期一", "2019-06-25", "gz")
        for i in range(2, max_row_num + 1):
            temp_list = []
            temp_list.append(self.sheet.cell(row=i, column=2).value)
            temp_list.append(self.sheet.cell(row=i, column=3).value)
            data_list.append(temp_list)
        return data_list


if __name__ == '__main__':
    ou = OpenpyxlUtil("./ceshishuju.xlsx", "Sheet1")
    d_list = ou.getDataFromSheet()
    for data in d_list:
        print(data)
