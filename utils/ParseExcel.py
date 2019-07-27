# -*- cocoding:utf-8 -*-
import openpyxl
from openpyxl.styles import Border, Side, Font
import time


class ParseExcel:
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)  # 设置字体颜色
        self.RGB_Dict = {"red": "FFFF3030", "green": "FF))8B00"}  # 颜色对应的RGB值

    def loadWorkBook(self, excelPathAndName):
        # 将Excel文件加载到内存，并获取其workbook对象
        try:
            self.excelFile = excelPathAndName
            print(excelPathAndName)
            self.workbook = openpyxl.load_workbook(excelPathAndName)
            return self.workbook
        except Exception as e:
            raise e

    def getSheetByName(self, sheetName):
        # 根据sheet名称获取sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        # 根据sheet的索引获取sheet对象
        try:
            sheet_name = self.workbook.get_sheet_names()[sheetIndex]
            sheet = self.workbook.get_sheet_by_name(sheet_name)
            return sheet
        except Exception as e:
            raise e

    def getRowsNumber(self, sheet):
        # 获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColumnNumber(self, sheet):
        # 获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取sheet中有数据区域的开始行号
        return sheet.min_row

    def getStartColumnNumber(self, sheet):
        # 获取sheet中有数据区域的开始列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple
        # 下标从1开始，sheet.rows[1]表示第一行
        try:
            return list(sheet.rows)[rowNo]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple
        # 下标从1开始，sheet.columns[1]表示第一列
        try:
            return list(sheet.columns)[colNo]
        except Exception as e:
            raise e

    def getCellObject(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 获取某个单元格的对象，可以直接根据Excel中单元格的编码及坐标，也可以根据单元格所在位置的数字索引
        # sheet.cell(coordinate="A1")或者sheet.cell(row=1, column=1)
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellValue(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 获取某个单元格中的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colNo=None, style=None):
        # 根据单元格在Excel中的编码坐标或数字索引坐标向单元格中写入数据
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGB_Dict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = content
                if style is not None:
                    sheet.cell(row=rowNo, column=colNo).font = Font(color=self.RGB_Dict[style])
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates if cell!")

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 写入当前时间
        now = int(time.time())  # 显示时间为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")


if __name__ == '__main__':
    pe = ParseExcel()
    pe.loadWorkBook("./测试.xlsx")
    # sheet = pe.getSheetByName("用例")
    sheet = pe.getSheetByIndex(0)
    print("行数：",pe.getRowsNumber(sheet))
    print("列数：",pe.getColumnNumber(sheet))
    row1=pe.getRow(sheet,2)
    for i in row1:
        print("第二行内容：", i.value)
    print(pe.getCellValue(sheet,rowNo=4,colNo=4))
    pe.writeCell(sheet,"test",rowNo=7,colNo=4,style="red")
    pe.writeCellCurrentTime(sheet,rowNo=7,colNo=5)