# -*- coding:utf-8 -*-
from openpyxl import load_workbook
import pysnooper


class ParseExcel:
    def __init__(self, excel_path, sheet_name):
        # 加载Excel数据
        self.wb = load_workbook(excel_path)
        # 加载sheet_name工作表数据
        self.ws = self.wb[sheet_name]  #workbook.get_sheet_by_name(sheet_name) 已弃用
        # 获取sheet_name工作表中的最大行号
        self.row_num = self.ws.max_row

    @pysnooper.snoop()
    def getDataFromSheet(self):
        data_list = []
        # 去掉第一行标题
        for row in range(2, self.row_num + 1):
            temp_list = []
            temp_list.append(self.ws.cell(row, 2).value)
            temp_list.append(self.ws.cell(row, 3).value)
            data_list.append(temp_list)
        return data_list


if __name__ == '__main__':
    pe = ParseExcel("./ceshishuju.xlsx", "Sheet1")
    d_list = pe.getDataFromSheet()
    for data in d_list:
        print(data)
